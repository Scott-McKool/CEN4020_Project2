import re
import sqlite3

###################################
### For accounts and logging in ###
###################################

def try_login(username, password) -> bool:

    conn = sqlite3.connect("users.db")
    cursor = conn.cursor()
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL
        )
    ''')

    cursor.execute('SELECT password FROM users WHERE username = ?', (username,))
    row = cursor.fetchone()
    conn.close()
    if row is None:
        return False
    
    passwd = row[0]

    if passwd == password:
        return True
    return False

#################################
### Course catalog data        ###
#################################

# [CRN, Name, Subject, CourseNum, Level (UG/GR), Instructor, Time, Room, Enrolled, Semester]
# Loaded from: Bellini Classes F25.xlsx, S25.xlsx, S26_NewClassesToBeAdded.xlsx
COURSES = [
    [91940, "Advanced Database Systems", "COP", "4703", "UG", "Fang, Suey-Chyun", "F 11:00 AM - 01:45 PM", "BEH 104", 99, "Fall 2025"],
    [92859, "Advanced OOP in C++", "CIS", "4930", "UG", "Jeanty, Henry H", "TR 09:30 AM - 10:45 AM", "ENB 118", 9, "Fall 2025"],
    [91938, "Advanced Program Design", "COP", "3515", "UG", "Holloman, Amanda K", "TR 05:00 PM - 06:15 PM", "ULH 101", 141, "Fall 2025"],
    [93316, "Advances in Data Structures", "COP", "6536", "GR", "Chakraburtty, Manisha ", "TR 02:00 PM - 03:15 PM", "ULH 101", 9, "Fall 2025"],
    [97869, "AI Bootcamp", "EGN", "4930", "UG", "Pamplona Segundo, Mauricio", "MTWRF 10:00 AM - 05:00 PM", "TBAT TBA", 14, "Fall 2025"],
    [97765, "AI for Integrated Sensing Comm", "CIS", "6930", "GR", "Zhang, Shichen", "TR 06:30 PM - 07:45 PM", "EDU 347", 5, "Fall 2025"],
    [97766, "AI for Integrated Sensing Comm", "CIS", "4930", "UG", "Zhang, Shichen", "TR 06:30 PM - 07:45 PM", "EDU 347", 9, "Fall 2025"],
    [91941, "Analysis Of Algorithms", "COT", "4400", "UG", "Topsakal, Oguzhan ", "TR 05:00 PM - 06:15 PM", "CWY 107", 162, "Fall 2025"],
    [97672, "Analysis Of Algorithms", "COT", "4400", "UG", "Qiao, Fengchun", "TR 05:00 PM - 06:15 PM", "EDU 347", 45, "Fall 2025"],
    [95430, "Augmented Reality", "CIS", "4930", "UG", "Han, Zhao ", "TR 08:00 AM - 09:15 AM", "SOC 150", 13, "Fall 2025"],
    [98256, "Augmented Reality", "CAP", "6110", "GR", "Han, Zhao ", "TR 08:00 AM - 09:15 AM", "SOC 150", 9, "Fall 2025"],
    [91945, "Automata Thry/Formal Languages", "COT", "4210", "UG", "Korzhova, Valentina N", "F 11:00 AM - 01:45 PM", "CWY 107", 61, "Fall 2025"],
    [93981, "Bellini College TA Training", "CIS", "6930", "GR", "Small, Schinnel K", "TBA", "TBAT TBA", 27, "Fall 2025"],
    [91394, "C Programming Essentials", "COP", "5227", "GR", "Gaspar, Alessio", "TBA", "OFFT OFF", 18, "Fall 2025"],
    [91947, "Capture the Flag", "CIS", "4930", "UG", "Pazos Revilla, Marbin", "MW 12:30 PM - 01:45 PM", "CHE 101", 40, "Fall 2025"],
    [91949, "Capture the Flag", "COP", "4931", "UG", "Alam, Md Rabbi", "MW 12:30 PM - 01:45 PM", "CPR 125", 39, "Fall 2025"],
    [93150, "Cloud Computing", "CIS", "6082", "GR", "Ventura, Philip R", "TR 03:30 PM - 04:45 PM", "SOC 149", 15, "Fall 2025"],
    [91951, "Cloud Computing for IT", "CIS", "4083", "UG", "Ventura, Philip R", "TR 03:30 PM - 04:45 PM", "SOC 149", 67, "Fall 2025"],
    [97679, "Cloud Computing for IT", "CIS", "4083", "UG", "Ventura, Philip R", "TR 11:00 AM - 12:15 PM", "BSN 1102", 45, "Fall 2025"],
    [91842, "CMOS-VLSI Design", "CIS", "6930", "GR", "Sheng, Yi", "MW 08:00 AM - 09:15 AM", "PCD 1147", 9, "Fall 2025"],
    [91952, "CMOS-VLSI Design", "CDA", "4213", "UG", "Sheng, Yi", "MW 08:00 AM - 09:15 AM", "PCD 1147", 67, "Fall 2025"],
    [92905, "CMOS-VLSI Design Lab", "CDA", "4213L", "UG", "Sheng, Yi", "T 10:00 AM - 11:45 AM", "ENB 214", 30, "Fall 2025"],
    [92906, "CMOS-VLSI Design Lab", "CDA", "4213L", "UG", "Sheng, Yi", "T 12:00 PM - 01:45 PM", "ENB 214", 36, "Fall 2025"],
    [97596, "Computer Architecture", "CDA", "4205", "UG", "Huang, Jiajun", "TR 02:00 PM - 03:15 PM", "CHE 101", 64, "Fall 2025"],
    [91955, "Computer Architecture", "CDA", "4205", "UG", "Karam, Robert A", "MW 03:30 PM - 04:45 PM", "CWY 109", 154, "Fall 2025"],
    [97824, "Computer Architecture", "CDA", "4205", "UG", "Pandey, Santosh", "MW 03:30 PM - 04:45 PM", "CHE 102", 38, "Fall 2025"],
    [92903, "Computer Architecture Lab", "CDA", "4205L", "UG", "Karam, Robert A", "W 08:00 AM - 09:45 AM", "ENB 214", 29, "Fall 2025"],
    [97659, "Computer Architecture Lab", "CDA", "4205L", "UG", "Pandey, Santosh", "W 06:00 PM - 07:45 PM", "ENB 214", 33, "Fall 2025"],
    [92900, "Computer Architecture Lab", "CDA", "4205L", "UG", "Huang, Jiajun", "R 08:00 AM - 09:45 AM", "ENB 214", 35, "Fall 2025"],
    [92901, "Computer Architecture Lab", "CDA", "4205L", "UG", "Huang, Jiajun", "R 12:00 PM - 01:45 PM", "ENB 214", 38, "Fall 2025"],
    [92902, "Computer Architecture Lab", "CDA", "4205L", "UG", "Karam, Robert A", "W 10:00 AM - 11:45 AM", "ENB 214", 37, "Fall 2025"],
    [92904, "Computer Architecture Lab", "CDA", "4205L", "UG", "Karam, Robert A", "W 12:00 PM - 01:45 PM", "ENB 214", 38, "Fall 2025"],
    [95881, "Computer Architecture Lab", "CDA", "4205L", "UG", "Pandey, Santosh", "R 10:00 AM - 11:45 AM", "ENB 214", 36, "Fall 2025"],
    [92870, "Computer Info Networks Lab", "CNT", "4104L", "UG", "Rauscher, Richard L", "R 02:00 PM - 03:45 PM", "NEC 300", 13, "Fall 2025"],
    [92878, "Computer Info Networks Lab", "CNT", "4104L", "UG", "Rauscher, Richard L", "R 10:00 AM - 11:45 AM", "NEC 300", 32, "Fall 2025"],
    [92882, "Computer Info Networks Lab", "CNT", "4104L", "UG", "Rauscher, Richard L", "R 12:00 PM - 01:45 PM", "NEC 300", 31, "Fall 2025"],
    [92884, "Computer Info Networks Lab", "CNT", "4104L", "UG", "Rauscher, Richard L", "F 10:00 AM - 11:45 AM", "NEC 300", 32, "Fall 2025"],
    [95436, "Computer Info Networks Lab", "CNT", "4104L", "UG", "Rauscher, Richard L", "F 12:00 PM - 01:45 PM", "NEC 300", 25, "Fall 2025"],
    [91973, "Computer Information Networks", "CNT", "4104", "UG", "Rauscher, Richard L", "TR 08:00 AM - 09:15 AM", "ULH 101", 137, "Fall 2025"],
    [91918, "Computer Logic & Design", "CDA", "3201", "UG", "Rahman, Taseef ", "TR 12:30 PM - 01:45 PM", "CWY 107", 109, "Fall 2025"],
    [97734, "Computer Logic & Design", "CDA", "3201", "UG", "Rahman, Taseef ", "TR 03:30 PM - 04:45 PM", "SOC 152", 29, "Fall 2025"],
    [92910, "Computer Logic Design Lab", "CDA", "3201L", "UG", "Rahman, Taseef ", "F 08:00 AM - 09:45 AM", "ENB 214", 17, "Fall 2025"],
    [93280, "Computer Logic Design Lab", "CDA", "3201L", "UG", "Rahman, Taseef ", "F 10:00 AM - 11:45 AM", "ENB 214", 26, "Fall 2025"],
    [93281, "Computer Logic Design Lab", "CDA", "3201L", "UG", "Rahman, Taseef ", "F 12:00 PM - 01:45 PM", "ENB 214", 22, "Fall 2025"],
    [92907, "Computer Logic Design Lab", "CDA", "3201L", "UG", "Rahman, Taseef ", "R 04:00 PM - 05:45 PM", "ENB 214", 38, "Fall 2025"],
    [92908, "Computer Logic Design Lab", "CDA", "3201L", "UG", "Rahman, Taseef ", "R 06:00 PM - 07:45 PM", "ENB 214", 36, "Fall 2025"],
    [91980, "Computer Organization", "CDA", "3103", "UG", "Rahman, Taseef ", "MW 12:30 PM - 01:45 PM", "BSN 1201", 79, "Fall 2025"],
    [91983, "Computer Organization", "CDA", "3103", "UG", "Zhang, Yan", "TR 12:30 PM - 01:45 PM", "CMC 141", 162, "Fall 2025"],
    [95874, "Computer Vision", "CAI", "4841", "UG", "Pamplona Segundo, Mauricio", "MW 12:30 PM - 01:45 PM", "CHE 102", 50, "Fall 2025"],
    [97805, "Computer Vision", "CAI", "5845", "GR", "Sarkar, Sudeep", "TBA", "OFFT OFF", 11, "Fall 2025"],
    [97806, "Computer Vision", "CAI", "5845", "GR", "Sarkar, Sudeep", "TBA", "OFFT OFF", 13, "Fall 2025"],
    [97850, "Computer Vision", "CAI", "4841", "UG", "Ojha, Utkarsh", "MW 11:00 AM - 12:15 PM", "BSN 1101", 35, "Fall 2025"],
    [91392, "Computing Essentials", "COP", "5008", "GR", "Anderson, James M", "TBA", "OFFT OFF", 19, "Fall 2025"],
    [97308, "Computing Essentials", "COP", "5008", "GR", "Anderson, James M", "TBA", "TBA", 0, "Fall 2025"],
    [97309, "Computing Essentials", "COP", "5008", "GR", "Anderson, James M", "TBA", "TBA", 7, "Fall 2025"],
    [97310, "Computing Essentials", "COP", "5008", "GR", "Anderson, James M", "T 06:30 PM - 08:05 PM", "TBA", 2, "Fall 2025"],
    [91844, "Computing Massive Parallel Sys", "COP", "6527", "GR", "Tu, Yicheng", "MW 09:30 AM - 10:45 AM", "BSN 1201", 13, "Fall 2025"],
    [91962, "Computing Massive Parallel Sys", "COP", "4520", "UG", "Tu, Yicheng", "MW 09:30 AM - 10:45 AM", "BSN 1201", 40, "Fall 2025"],
    [91877, "Control of Mobile Robots", "CDA", "4621", "UG", "Weitzenfeld, Alfredo", "TR 03:30 PM - 04:45 PM", "CPR 115", 39, "Fall 2025"],
    [91838, "Cryptographic Hardware", "CDA", "6328", "GR", "Mozaffari Kermani, Mehran", "MW 11:00 AM - 12:15 PM", "BSN 1201", 40, "Fall 2025"],
    [91988, "Cryptographic Hardware", "CDA", "4321", "UG", "Mozaffari Kermani, Mehran", "MW 11:00 AM - 12:15 PM", "BSN 1201", 45, "Fall 2025"],
    [97617, "Cryptography Theory & Practice", "CIS", "6930", "GR", "Yavuz, Attila A", "TR 02:00 PM - 03:15 PM", "ENB 118", 10, "Fall 2025"],
    [97618, "Cryptography Theory & Practice", "CIS", "4930", "UG", "Yavuz, Attila A", "TR 02:00 PM - 03:15 PM", "ENB 118", 40, "Fall 2025"],
    [91990, "CSE Project", "CIS", "4910", "UG", "Hollingsworth, Scot A", "F 11:00 AM - 01:45 PM", "ENG 4", 32, "Fall 2025"],
    [91992, "Data Stg & Anlys with Hadoop", "CIS", "4345", "UG", "Zhang, Yan", "MW 09:30 AM - 10:45 AM", "SOC 152", 74, "Fall 2025"],
    [91994, "Data Structures", "COP", "4530", "UG", "Korzhova, Valentina N", "TR 02:00 PM - 03:15 PM", "CPR 103", 167, "Fall 2025"],
    [91878, "Data Structures & Algorithms", "COP", "4538", "UG", "Chakraburtty, Manisha ", "TR 02:00 PM - 03:15 PM", "ULH 101", 145, "Fall 2025"],
    [91395, "Data Structures Essentials", "COP", "5532", "GR", "Hidalgo, Isabela C", "TBA", "OFFT OFF", 16, "Fall 2025"],
    [92072, "Data Structures Essentials", "COP", "5532", "GR", "Hidalgo, Isabela C", "TBA", "OFFT OFF", 14, "Fall 2025"],
    [97314, "Data Structures Essentials", "COP", "5532", "GR", "Hidalgo, Isabela C", "TBA", "OFFT OFF", 6, "Fall 2025"],
    [97315, "Data Structures Essentials", "COP", "5532", "GR", "Hidalgo, Isabela C", "R 06:30 PM - 08:05 PM", "TBA", 3, "Fall 2025"],
    [91996, "Database Design", "COP", "4710", "UG", "Fang, Suey-Chyun", "MW 05:00 PM - 06:15 PM", "CHE 103", 43, "Fall 2025"],
    [97424, "Digital Image Processing", "CAI", "5846", "GR", "Goldgof, Dmitry B", "MW 03:30 PM - 04:45 PM", "CHE 303", 27, "Fall 2025"],
    [91390, "Discrete Structures Essentials", "COT", "5105", "GR", "Wang, Jing", "TBA", "OFFT OFF", 17, "Fall 2025"],
    [97306, "Discrete Structures Essentials", "COT", "5105", "GR", "Wang, Jing", "TBA", "TBA", 6, "Fall 2025"],
    [97307, "Discrete Structures Essentials", "COT", "5105", "GR", "Wang, Jing", "R 06:30 PM - 08:05 PM", "TBA", 3, "Fall 2025"],
    [91998, "Ethical Issues/Profess Conduct", "CIS", "4250", "UG", "Murray-Bruce, Oluwaremilekun John", "MW 03:30 PM - 04:45 PM", "CHE 101", 63, "Fall 2025"],
    [91997, "Ethical Issues/Profess Conduct", "CIS", "4250", "UG", "Baten, Raiyan Abdul ", "TR 06:30 PM - 07:45 PM", "CWY 107", 56, "Fall 2025"],
    [92002, "Foundations of Cybersecurity", "CIS", "3213", "UG", "Ren, Yili ", "TBA", "TBAT TBA", 62, "Fall 2025"],
    [91999, "Foundations of Cybersecurity", "CIS", "3213", "UG", "Gauvin Jr, William J", "TBA", "TBAT TBA", 91, "Fall 2025"],
    [97494, "Freshman Seminar", "CIS", "1930", "UG", "Morgan, John P", "TBA", "OFFT OFF", 398, "Fall 2025"],
    [97964, "Generative AI", "CIS", "6930", "GR", "Karimian, Nima", "TR 11:00 AM - 12:15 PM", "CPR 120", 19, "Fall 2025"],
    [97965, "Generative AI", "CIS", "4930", "UG", "Karimian, Nima", "TR 11:00 AM - 12:15 PM", "CPR 120", 28, "Fall 2025"],
    [95286, "Hands-on Cybersecurity", "CIS", "4622", "UG", "Alam, Md Rabbi", "TR 03:30 PM - 04:45 PM", "ENB 216", 49, "Fall 2025"],
    [92004, "Hands-on Cybersecurity", "CIS", "4622", "UG", "Ou, Xinming", "TR 12:30 PM - 01:45 PM", "ENB 216", 54, "Fall 2025"],
    [92005, "Hands-on Cybersecurity", "CIS", "4622", "UG", "Fang, Suey-Chyun", "MW 11:00 AM - 12:15 PM", "ENB 216", 42, "Fall 2025"],
    [91843, "Human Aspects of Cybersecurity", "CIS", "6218", "GR", "Caulkins, Bruce D", "MW 03:30 PM - 04:45 PM", "TBAT TBA", 0, "Fall 2025"],
    [92006, "Human Aspects of Cybersecurity", "CIS", "4219", "UG", "Caulkins, Bruce D", "MW 03:30 PM - 04:45 PM", "TBAT TBA", 93, "Fall 2025"],
    [92007, "Human-Computer Interaction", "CIS", "4930", "UG", "Woodward, Julia E", "TR 09:30 AM - 10:45 AM", "CHE 103", 18, "Fall 2025"],
    [98156, "Human-Computer Interaction", "CAP", "5178", "GR", "Woodward, Julia E", "TR 09:30 AM - 10:45 AM", "CHE 103", 50, "Fall 2025"],
    [97631, "Incident Response Methods", "COP", "4931", "UG", "Irving, Ryan ", "TBA", "TBAT TBA", 77, "Fall 2025"],
    [91915, "Intro to Computer Programming", "CGS", "2060", "UG", "Holloman, Amanda K", "MW 05:00 PM - 06:15 PM", "BEH 103", 70, "Fall 2025"],
    [91840, "Intro to Theory of Algorithms", "COT", "6405", "GR", "Canavan, Shaun", "TR 05:00 PM - 06:15 PM", "PCD 1147", 85, "Fall 2025"],
    [95471, "Introduction to AI", "CAI", "5005", "GR", "Sun, Yu", "TR 05:00 PM - 06:15 PM", "CHE 101", 30, "Fall 2025"],
    [97552, "Introduction to AI", "CAI", "5005", "GR", "Sun, Yu", "TR 05:00 PM - 06:15 PM", "CHE 101", 15, "Fall 2025"],
    [95467, "Introduction to AI", "CAI", "4002", "UG", "Kim, Gene L", "TR 06:30 PM - 07:45 PM", "BEH 103", 82, "Fall 2025"],
    [95962, "Introduction to AI", "CAI", "5005", "GR", "Pamplona Segundo, Mauricio", "TBA", "OFFT OFF", 16, "Fall 2025"],
    [95963, "Introduction to AI", "CAI", "5005", "GR", "Pamplona Segundo, Mauricio", "TBA", "OFFT OFF", 3, "Fall 2025"],
    [97678, "Introduction to AI", "CAI", "4002", "UG", "Pamplona Segundo, Mauricio", "F 08:00 AM - 10:45 AM", "CHE 100", 42, "Fall 2025"],
    [97919, "Introduction to AI", "CAI", "5005", "GR", "Pamplona Segundo, Mauricio", "TBA", "OFFT OFF", 15, "Fall 2025"],
    [91939, "Introduction to Databases", "CGS", "1540", "UG", "Hidalgo, Isabela C", "MW 09:30 AM - 10:45 AM", "BEH 104", 121, "Fall 2025"],
    [97821, "Introduction to Databases", "CGS", "1540", "UG", "Ventura, Philip R", "MW 11:00 AM - 12:15 PM", "SOC 145", 43, "Fall 2025"],
    [97822, "Introduction to Databases", "CGS", "1540", "UG", "Ventura, Philip R", "MW 12:30 PM - 01:45 PM", "SOC 132", 43, "Fall 2025"],
    [91942, "Introduction to Robotics", "CAP", "4662", "UG", "Sun, Yu", "TR 03:30 PM - 04:45 PM", "CHE 103", 41, "Fall 2025"],
    [97904, "Introduction to Robotics", "CIS", "6930", "GR", "Sun, Yu", "TR 03:30 PM - 04:45 PM", "CHE 103", 2, "Fall 2025"],
    [97755, "Machine Learning", "CAI", "5107", "GR", "Kabir, Anowarul", "TR 03:30 PM - 04:45 PM", "CHE 101", 47, "Fall 2025"],
    [97858, "Math for AI", "CAI", "5035", "GR", "Pamplona Segundo, Mauricio", "TBA", "TBA", 4, "Fall 2025"],
    [97859, "Math for AI", "CAI", "5035", "GR", "Pamplona Segundo, Mauricio", "TBA", "OFFT OFF", 4, "Fall 2025"],
    [94947, "Mobile Biometrics", "CAP", "4103", "UG", "Neal, Tempestt J", "TR 08:00 AM - 09:15 AM", "EDU 347", 39, "Fall 2025"],
    [94948, "Mobile Biometrics", "CAP", "6101", "GR", "Neal, Tempestt J", "TR 08:00 AM - 09:15 AM", "EDU 347", 31, "Fall 2025"],
    [98258, "Net Analysis & ML with Graphs", "CAI", "5155", "GR", "Kim, Seungbae ", "F 11:00 AM - 01:45 PM", "CHE 101", 26, "Fall 2025"],
    [95474, "Network Analysis & ML", "CIS", "4930", "UG", "Kim, Seungbae ", "F 11:00 AM - 01:45 PM", "CHE 101", 30, "Fall 2025"],
    [91953, "Network Security & Firewalls", "CNT", "4403", "UG", "Gauvin Jr, William J", "TBA", "TBAT TBA", 126, "Fall 2025"],
    [91882, "Object Oriented Programming", "COP", "2513", "UG", "Hidalgo, Isabela C", "TR 09:30 AM - 10:45 AM", "CHE 100", 188, "Fall 2025"],
    [97820, "Operating Systems", "COP", "6611", "GR", "Ahmed, Shareef", "TR 12:30 PM - 01:45 PM", "CIS 1047", 27, "Fall 2025"],
    [91954, "Operating Systems", "COP", "4600", "UG", "Yi, Hye S", "MW 02:00 PM - 03:15 PM", "CPR 115", 95, "Fall 2025"],
    [91831, "Operating Systems", "COP", "6611", "GR", "Templeton, John ", "MW 05:00 PM - 06:15 PM", "CIS 1048", 83, "Fall 2025"],
    [91957, "Operating Systems", "COP", "4600", "UG", "Yi, Hye S", "MW 11:00 AM - 12:15 PM", "CPR 115", 89, "Fall 2025"],
    [97605, "Operating Systems", "COP", "4600", "UG", "Zhang, Yan", "MW 02:00 PM - 03:15 PM", "SOC 152", 82, "Fall 2025"],
    [91958, "Penetration Testing", "CIS", "4200", "UG", "Kafle, Kaushal ", "MW 02:00 PM - 03:15 PM", "FAH 101", 106, "Fall 2025"],
    [95473, "Penetration Testing for IT", "CIS", "6220", "GR", "Kafle, Kaushal ", "MW 02:00 PM - 03:15 PM", "FAH 101", 1, "Fall 2025"],
    [91832, "Princ of Computer Architecture", "EEL", "6764", "GR", "Zheng, Hao", "TR 06:30 PM - 07:45 PM", "CPR 103", 48, "Fall 2025"],
    [91961, "Program Design", "COP", "3514", "UG", "Wang, Jing", "TR 09:30 AM - 10:45 AM", "CWY 109", 180, "Fall 2025"],
    [91960, "Program Design", "COP", "3514", "UG", "Wang, Jing", "MW 03:30 PM - 04:45 PM", "EDU 347", 71, "Fall 2025"],
    [92947, "Programming Concepts", "COP", "2510", "UG", "Kaleemunnisa, Fnu", "TR 05:00 PM - 06:15 PM", "CPR 103", 40, "Fall 2025"],
    [92948, "Programming Concepts", "COP", "2510", "UG", "Kaleemunnisa, Fnu", "TR 05:00 PM - 06:15 PM", "CPR 103", 40, "Fall 2025"],
    [92949, "Programming Concepts", "COP", "2510", "UG", "Kaleemunnisa, Fnu", "TR 05:00 PM - 06:15 PM", "CPR 103", 40, "Fall 2025"],
    [92951, "Programming Concepts", "COP", "2510", "UG", "Small, Schinnel K", "TR 05:00 PM - 06:15 PM", "CIS 1048", 40, "Fall 2025"],
    [92952, "Programming Concepts", "COP", "2510", "UG", "Kaleemunnisa, Fnu", "TR 05:00 PM - 06:15 PM", "CPR 103", 40, "Fall 2025"],
    [92954, "Programming Concepts", "COP", "2510", "UG", "Small, Schinnel K", "TR 05:00 PM - 06:15 PM", "CIS 1048", 40, "Fall 2025"],
    [97750, "Programming Concepts", "COP", "2510", "UG", "Small, Schinnel K", "TR 05:00 PM - 06:15 PM", "CIS 1048", 40, "Fall 2025"],
    [97751, "Programming Concepts", "COP", "2510", "UG", "Yi, Hye S", "TR 08:00 AM - 09:15 AM", "CHE 111", 40, "Fall 2025"],
    [97752, "Programming Concepts", "COP", "2510", "UG", "Yi, Hye S", "TR 08:00 AM - 09:15 AM", "CHE 111", 40, "Fall 2025"],
    [97753, "Programming Concepts", "COP", "2510", "UG", "Yi, Hye S", "TR 08:00 AM - 09:15 AM", "CHE 111", 40, "Fall 2025"],
    [97960, "Programming Concepts", "COP", "2510", "UG", "Small, Schinnel K", "TR 05:00 PM - 06:15 PM", "CIS 1048", 32, "Fall 2025"],
    [97979, "Programming Concepts", "COP", "2510", "UG", "Yi, Hye S", "TR 08:00 AM - 09:15 AM", "CHE 111", 40, "Fall 2025"],
    [98248, "Programming Concepts", "COP", "2510", "UG", "Kaleemunnisa, Fnu", "TR 05:00 PM - 06:15 PM", "TBA", 25, "Fall 2025"],
    [98249, "Programming Concepts", "COP", "2510", "UG", "Kaleemunnisa, Fnu", "TR 05:00 PM - 06:15 PM", "TBA", 40, "Fall 2025"],
    [91964, "Secure Coding", "CNT", "4419", "UG", "Liu, Yao", "MW 03:30 PM - 04:45 PM", "CWY 107", 110, "Fall 2025"],
    [97674, "Secure Coding", "CNT", "4419", "UG", "Ami, Amit Seal", "TR 09:30 AM - 10:45 AM", "CPR 127", 41, "Fall 2025"],
    [91835, "Seminar in AI", "CIS", "6930", "GR", "Hall, Lawrence O", "F 01:00 PM - 01:50 PM", "ENB 118", 10, "Fall 2025"],
    [91968, "Software Engineering", "CEN", "4020", "UG", "Anderson, James M", "MW 05:00 PM - 06:15 PM", "CWY 107", 185, "Fall 2025"],
    [92857, "Software Sys Dev for CYS", "COP", "4931", "UG", "Jeanty, Henry H", "MW 12:30 PM - 01:45 PM", "ENB 118", 22, "Fall 2025"],
    [92861, "Software System Development", "COP", "4365", "UG", "Jeanty, Henry H", "TR 12:30 PM - 01:45 PM", "ENB 118", 73, "Fall 2025"],
    [94945, "Software System Development", "COP", "4365", "UG", "Jeanty, Henry H", "MW 09:30 AM - 10:45 AM", "ENB 118", 36, "Fall 2025"],
    [91970, "System Admin & Maintenance", "CNT", "4603", "UG", "Hoeke, Marilynn Carol", "TBA", "TBAT TBA", 68, "Fall 2025"],
    [93038, "Topics in NLP", "CIS", "6930", "GR", "Chhabra, Anshuman ", "TR 08:00 AM - 09:15 AM", "BSN 1201", 40, "Fall 2025"],
    [98257, "Trustworthy AI Systems", "CAI", "6605", "GR", "Wang, Guangjing ", "MW 06:30 PM - 07:45 PM", "CHE 103", 45, "Fall 2025"],
    [95920, "User Experience Design", "CIS", "4930", "UG", "Holloman, Amanda K", "MW 12:30 PM - 01:45 PM", "CHE 303", 60, "Fall 2025"],
    [97445, "VLSI Testing", "CIS", "6930", "GR", "Zheng, Hao", "MW 11:00 AM - 12:15 PM", "SOC 149", 7, "Fall 2025"],
    [97619, "VLSI Testing", "CIS", "4930", "UG", "Zheng, Hao", "MW 11:00 AM - 12:15 PM", "SOC 149", 28, "Fall 2025"],
    [91971, "Web Systems for IT", "CGS", "3853", "UG", "Gaspar, Alessio", "MW 09:30 AM - 10:45 AM", "CHE 103", 78, "Fall 2025"],
    [15533, "Adv Database Systems for IT", "COP", "4703", "UG", "Suey-Chyun Fang", "F 11:00 AM - 01:45 PM", "CWY 107", 0, "Spring 2025"],
    [15509, "Adv Program Design", "COP", "3515", "UG", "Hye Yi", "MW 12:30 PM - 01:45 PM", "CMC 141", 0, "Spring 2025"],
    [15541, "Advanced OOP in C++", "CIS", "4930", "UG", "Henry Jeanty", "TR 12:30 PM - 01:45 PM", "CHE 103", 0, "Spring 2025"],
    [15542, "Advanced OOP in C++ for IT", "COP", "4931", "UG", "Henry Jeanty", "MW 12:30 PM - 01:45 PM", "ENB 118", 0, "Spring 2025"],
    [19011, "Advances in Data Structures", "COP", "6536", "GR", "Philip Ventura", "TR 02:00 PM - 03:15 PM", "CPR 115", 0, "Spring 2025"],
    [18954, "Affective Computing", "CAI", "5615", "GR", "Shaun Canavan", "MW 05:00 PM - 06:15 PM", "CPR 103", 0, "Spring 2025"],
    [15933, "Affective Computing", "CAP", "4628", "UG", "Shaun Canavan", "MW 05:00 PM - 06:15 PM", "CPR 103", 0, "Spring 2025"],
    [15510, "Analysis Of Algorithms", "COT", "4400", "UG", "Oguzhan Topsakal", "MW 05:00 PM - 06:15 PM", "ISA 1061", 0, "Spring 2025"],
    [15511, "Analysis Of Algorithms", "COT", "4400", "UG", "Oguzhan Topsakal", "MW 11:00 AM - 12:15 PM", "CHE 111", 0, "Spring 2025"],
    [15506, "Automata Thry/Formal Languages", "COT", "4210", "UG", "Valentina Korzhova", "F 11:00 AM - 01:45 PM", "CHE 100", 0, "Spring 2025"],
    [19021, "Autonomous Mobile Robots", "CAI", "5815", "GR", "Alfredo Weitzenfeld", "TR 03:30 PM - 04:45 PM", "CHE 303", 0, "Spring 2025"],
    [18957, "Capture the Flag", "CIS", "4930", "UG", "Marbin Pazos Revilla", "MW 11:00 AM - 12:15 PM", "BSF 100", 0, "Spring 2025"],
    [18958, "Capture the Flag", "COP", "4931", "UG", "Marbin Pazos Revilla", "MW 11:00 AM - 12:15 PM", "BSF 100", 0, "Spring 2025"],
    [15922, "Comp Methods for Imaging", "CIS", "4930", "UG", "Oluwaremilekun John Murray-Bruce", "TR 03:30 PM - 04:45 PM", "BEH 103", 0, "Spring 2025"],
    [15921, "Comp Methods for Imaging", "CIS", "6930", "GR", "Oluwaremilekun John Murray-Bruce", "TR 03:30 PM - 04:45 PM", "BEH 103", 0, "Spring 2025"],
    [15919, "Compilers", "COP", "4620", "UG", "Jarred Ligatti", "MW 05:00 PM - 06:15 PM", "CHE 100", 0, "Spring 2025"],
    [15920, "Compilers", "COP", "6625", "GR", "Jarred Ligatti", "MW 05:00 PM - 06:15 PM", "CHE 100", 0, "Spring 2025"],
    [15500, "Computer Architecture", "CDA", "4205", "UG", "Robert Karam", "MW 09:30 AM - 10:45 AM", "CPR 115", 0, "Spring 2025"],
    [15499, "Computer Architecture", "CDA", "4205", "UG", "Marvin Andujar", "TR 08:00 AM - 09:15 AM", "CMC 141", 0, "Spring 2025"],
    [16768, "Computer Architecture Lab", "CDA", "4205L", "UG", "Robert Karam", "F 09:30 AM - 11:15 AM", "ENB 214", 0, "Spring 2025"],
    [16769, "Computer Architecture Lab", "CDA", "4205L", "UG", "Robert Karam", "F 11:30 AM - 01:15 PM", "ENB 214", 0, "Spring 2025"],
    [16773, "Computer Architecture Lab", "CDA", "4205L", "UG", "Andujar, Marvin", "W 11:30 AM - 01:15 PM", "ENB 214", 0, "Spring 2025"],
    [16774, "Computer Architecture Lab", "CDA", "4205L", "UG", "Andujar, Marvin", "W 01:30 PM - 03:15 PM", "ENB 214", 0, "Spring 2025"],
    [16775, "Computer Architecture Lab", "CDA", "4205L", "UG", "Andujar, Marvin", "R 09:00 AM - 10:50 AM", "ENB 214", 0, "Spring 2025"],
    [16776, "Computer Architecture Lab", "CDA", "4205L", "UG", "Andujar, Marvin", "R 11:30 AM - 01:15 PM", "ENB 214", 0, "Spring 2025"],
    [16700, "Computer Info Networks for IT", "CNT", "4104", "UG", "Rauscher, Richard L", "TR 05:00 PM - 06:15 PM", "BSF 100", 0, "Spring 2025"],
    [15512, "Computer Logic & Design", "CDA", "3201", "UG", "Rahman, Taseef ", "TR 02:00 PM - 03:15 PM", "ULH 101", 0, "Spring 2025"],
    [15513, "Computer Logic & Design", "CDA", "3201", "UG", "Rahman, Taseef ", "MW 05:00 PM - 06:15 PM", "ULH 101", 0, "Spring 2025"],
    [16782, "Computer Logic Design Lab", "CDA", "3201L", "UG", "", "F 12:00 PM - 01:50 PM", "TBA", 0, "Spring 2025"],
    [16783, "Computer Logic Design Lab", "CDA", "3201L", "UG", "", "F 02:00 PM - 03:50 PM", "ENB 214", 0, "Spring 2025"],
    [16784, "Computer Logic Design Lab", "CDA", "3201L", "UG", "", "R 03:00 PM - 04:50 PM", "TBA", 0, "Spring 2025"],
    [16777, "Computer Logic Design Lab", "CDA", "3201L", "UG", "", "W 05:00 PM - 06:50 PM", "ENB 214", 0, "Spring 2025"],
    [16778, "Computer Logic Design Lab", "CDA", "3201L", "UG", "", "T 02:00 PM - 03:50 PM", "TBA", 0, "Spring 2025"],
    [16779, "Computer Logic Design Lab", "CDA", "3201L", "UG", "", "T 04:00 PM - 05:50 PM", "TBA", 0, "Spring 2025"],
    [16780, "Computer Logic Design Lab", "CDA", "3201L", "UG", "", "T 10:00 AM - 11:50 AM", "ENB 214", 0, "Spring 2025"],
    [16781, "Computer Logic Design Lab", "CDA", "3201L", "UG", "", "T 12:00 PM - 01:50 PM", "TBA", 0, "Spring 2025"],
    [20487, "Computer Networks I", "CNT", "4004", "UG", "", "TR 05:00 PM - 06:15 PM", "BSF 100", 0, "Spring 2025"],
    [16791, "Computer Networks Lab for IT", "CNT", "4104L", "UG", "", "R 09:00 AM - 10:50 AM", "NEC 300", 0, "Spring 2025"],
    [16790, "Computer Networks Lab for IT", "CNT", "4104L", "UG", "Rauscher, Richard L", "R 11:00 AM - 12:50 PM", "NEC 300", 0, "Spring 2025"],
    [16788, "Computer Networks Lab for IT", "CNT", "4104L", "UG", "Rauscher, Richard L", "R 01:00 PM - 02:50 PM", "NEC 300", 0, "Spring 2025"],
    [16789, "Computer Networks Lab for IT", "CNT", "4104L", "UG", "Rauscher, Richard L", "F 09:00 AM - 10:50 AM", "NEC 300", 0, "Spring 2025"],
    [16787, "Computer Networks Lab for IT", "CNT", "4104L", "UG", "Rauscher, Richard L", "R 03:00 PM - 04:50 PM", "NEC 300", 0, "Spring 2025"],
    [19006, "Computer Networks Lab for IT", "CNT", "4104L", "UG", "Rauscher, Richard L", "F 11:00 AM - 12:50 PM", "NEC 300", 0, "Spring 2025"],
    [15498, "Computer Organization", "CDA", "3103", "UG", "Yan Zhang", "MW 02:00 PM - 03:15 PM", "BSF 100", 0, "Spring 2025"],
    [15515, "Computer Organization", "CDA", "3103", "UG", "Yan Zhang", "TR 02:00 PM - 03:15 PM", "CWY 109", 0, "Spring 2025"],
    [16585, "Computer System Design", "CDA", "4203", "UG", "Mehran Mozaffari Kermani", "MW 03:30 PM - 04:45 PM", "CHE 217", 0, "Spring 2025"],
    [16785, "Computer System Design Lab", "CDA", "4203L", "UG", "", "T 01:30 PM - 03:15 PM", "ENB 214", 0, "Spring 2025"],
    [16786, "Computer System Design Lab", "CDA", "4203L", "UG", "", "T 03:30 PM - 05:15 PM", "ENB 214", 0, "Spring 2025"],
    [18996, "Computing Massive Parallel Sys", "COP", "4520", "UG", "Tu, Yicheng", "MW 09:30 AM - 10:45 AM", "SOC 149", 0, "Spring 2025"],
    [18997, "Computing Massive Parallel Sys", "COP", "6527", "GR", "Tu, Yicheng", "MW 09:30 AM - 10:45 AM", "SOC 149", 0, "Spring 2025"],
    [15704, "Control of Mobile Robots", "CDA", "4621", "UG", "Weitzenfeld, Alfredo", "TR 03:30 PM - 04:45 PM", "CHE 303", 0, "Spring 2025"],
    [16686, "CSE Project", "CIS", "4910", "UG", "Kenneth Christensen", "F 11:00 AM - 01:45 PM", "ENC 1002", 0, "Spring 2025"],
    [16792, "CyS Biometrics", "COP", "4931", "UG", "Neal, Tempestt J", "TR 09:30 AM - 10:45 AM", "BEH 104", 0, "Spring 2025"],
    [15516, "Data Structures", "COP", "4530", "UG", "Valentina Korzhova", "TR 02:00 PM - 03:15 PM", "FAH 101", 0, "Spring 2025"],
    [15517, "Data Structures", "COP", "4530", "UG", "Taseef Rahman", "MW 02:00 PM - 03:15 PM", "SOC 152", 0, "Spring 2025"],
    [15534, "Database Design", "COP", "4710", "UG", "Oguzhan Topsakal", "MW 02:00 PM - 03:15 PM", "CHE 111", 0, "Spring 2025"],
    [18998, "Deep Learning", "CAI", "5205", "GR", "Yu Sun", "TR 05:00 PM - 06:15 PM", "CHE 101", 0, "Spring 2025"],
    [16698, "Emrg Topics Network Security", "CNT", "6410", "GR", "Yao Liu", "MW 03:30 PM - 04:45 PM", "SOC 152", 0, "Spring 2025"],
    [16675, "Ethical Issues/Profess Conduct", "CIS", "4250", "UG", "John Licato", "TR 03:30 PM - 04:45 PM", "EDU 115", 0, "Spring 2025"],
    [16676, "Ethical Issues/Profess Conduct", "CIS", "4250", "UG", "Julia Woodward", "MW 05:00 PM - 06:15 PM", "CWY 107", 0, "Spring 2025"],
    [19926, "Ethical Issues/Profess Conduct", "CIS", "4250", "UG", "Amanda Holloman", "MW 12:30 PM - 01:45 PM", "BSN 1301", 0, "Spring 2025"],
    [18999, "Fair ML", "CIS", "6930", "GR", "Anshuman Chhabra", "MW 02:00 PM - 03:15 PM", "BSN 1201", 0, "Spring 2025"],
    [19000, "Foundations of Cybersecurity", "CIS", "3213", "UG", "William Gauvin Jr", "TBA", "OFFT OFF", 0, "Spring 2025"],
    [16691, "Hands-on Cybersecurity", "CIS", "4622", "UG", "Marbin Pazos Revilla", "TR 03:30 PM - 04:45 PM", "ENB 216", 0, "Spring 2025"],
    [16697, "Hands-on Cybersecurity", "CIS", "4622", "UG", "Philip Ventura", "TR 09:30 AM - 10:45 AM", "ENB 216", 0, "Spring 2025"],
    [15697, "Hardware Accelerators for ML", "CIS", "4930", "UG", "Dayane Alfenas Reis", "TR 09:30 AM - 10:45 AM", "BSN 1301", 0, "Spring 2025"],
    [15698, "Hardware Accelerators for ML", "CIS", "6930", "GR", "Dayane Alfenas Reis", "TR 09:30 AM - 10:45 AM", "BSN 1301", 0, "Spring 2025"],
    [15519, "Human Aspects of Cybersecurity", "CIS", "4219", "UG", "NING WANG", "TR 12:30 PM - 01:45 PM", "CIS 1048", 0, "Spring 2025"],
    [19005, "Human Aspects of Cybersecurity", "CIS", "4219", "UG", "", "TR 06:30 PM - 07:45 PM", "CMC 141", 0, "Spring 2025"],
    [19004, "Human Aspects of Cybersecurity", "CIS", "6218", "GR", "NING WANG", "TR 12:30 PM - 01:45 PM", "CIS 1048", 0, "Spring 2025"],
    [15520, "Intro to Computer Programming", "CGS", "2060", "UG", "James Anderson", "MW 05:00 PM - 06:15 PM", "BEH 104", 0, "Spring 2025"],
    [15521, "Intro to Databases for IT", "CGS", "1540", "UG", "Isabela Hidalgo", "MW 12:30 PM - 01:45 PM", "BEH 104", 0, "Spring 2025"],
    [15522, "Intro to Theory of Algorithms", "COT", "6405", "GR", "Seungbae Kim", "MW 11:00 AM - 12:15 PM", "CHE 103", 0, "Spring 2025"],
    [19010, "Introduction to AI", "CAI", "4002", "UG", "John Licato", "MW 03:30 PM - 04:45 PM", "SOC 150", 0, "Spring 2025"],
    [20335, "IoT Security", "CIS", "6930", "GR", "Kaushal Kafle", "F 11:00 AM - 01:45 PM", "BSN 1201", 0, "Spring 2025"],
    [15523, "IT Data Structures", "COP", "4538", "UG", "Philip Ventura", "TR 02:00 PM - 03:15 PM", "CPR 115", 0, "Spring 2025"],
    [15524, "IT Data Structures", "COP", "4538", "UG", "", "TR 11:00 AM - 12:15 PM", "CPR 115", 0, "Spring 2025"],
    [15501, "IT Object Oriented Programming", "COP", "2513", "UG", "Isabela Hidalgo", "MW 02:00 PM - 03:15 PM", "CWY 107", 0, "Spring 2025"],
    [15503, "IT Programming Fundamentals", "COP", "2512", "UG", "Alessio Gaspar", "MW 06:30 PM - 07:45 PM", "CWY 109", 0, "Spring 2025"],
    [15518, "Linux Command Line Interface", "COP", "4931", "UG", "Alessio Gaspar", "TR 05:00 PM - 06:15 PM", "SOC 151", 0, "Spring 2025"],
    [19012, "Natural Language Processing", "CAI", "5307", "GR", "Ankur Mali", "TR 06:30 PM - 07:45 PM", "CHE 101", 0, "Spring 2025"],
    [15700, "Natural Language Processing", "CAP", "4641", "UG", "Ankur Mali", "TR 06:30 PM - 07:45 PM", "CHE 101", 0, "Spring 2025"],
    [16696, "Network Security & Firewalls", "CNT", "4403", "UG", "William Gauvin Jr", "TBA", "OFFT OFF", 0, "Spring 2025"],
    [15526, "Operating Systems", "COP", "4600", "UG", "Mauricio Pamplona Segundo", "TR 06:30 PM - 07:45 PM", "CHE 103", 0, "Spring 2025"],
    [15527, "Operating Systems", "COP", "4600", "UG", "Mauricio Pamplona Segundo", "TR 06:30 PM - 07:45 PM", "CPR 115", 0, "Spring 2025"],
    [19013, "Operating Systems", "COP", "4600", "UG", "Hye Yi", "TR 09:30 AM - 10:45 AM", "SOC 151", 0, "Spring 2025"],
    [15507, "Operating Systems", "COP", "6611", "GR", "Gene Kim", "TR 05:00 PM - 06:15 PM", "CPR 115", 0, "Spring 2025"],
    [15925, "Penetration Testing for IT", "CIS", "4200", "UG", "Philip Ventura", "MW 08:00 AM - 09:15 AM", "FAH 101", 0, "Spring 2025"],
    [15926, "Penetration Testing for IT", "CIS", "6220", "GR", "Philip Ventura", "MW 08:00 AM - 09:15 AM", "FAH 101", 0, "Spring 2025"],
    [15508, "Princ of Computer Architecture", "EEL", "6764", "GR", "Srinivas Katkoori", "MW 05:00 PM - 06:15 PM", "CWY 109", 0, "Spring 2025"],
    [15497, "Program Design", "COP", "3514", "UG", "Jing Wang", "TR 12:30 PM - 01:45 PM", "CWY 107", 0, "Spring 2025"],
    [15505, "Program Design", "COP", "3514", "UG", "Jing Wang", "MW 06:30 PM - 07:45 PM", "CHE 100", 0, "Spring 2025"],
    [15745, "Programming Concepts", "COP", "2510", "UG", "Schinnel Small", "TR 03:30 PM - 04:45 PM", "BSN 1100", 0, "Spring 2025"],
    [15747, "Programming Concepts", "COP", "2510", "UG", "Schinnel Small", "TR 03:30 PM - 04:45 PM", "BSN 1100", 0, "Spring 2025"],
    [15746, "Programming Concepts", "COP", "2510", "UG", "Schinnel Small", "TR 03:30 PM - 04:45 PM", "BSN 1100", 0, "Spring 2025"],
    [15750, "Programming Concepts", "COP", "2510", "UG", "Schinnel Small", "TR 03:30 PM - 04:45 PM", "BSN 1100", 0, "Spring 2025"],
    [15753, "Programming Concepts", "COP", "2510", "UG", "Schinnel Small", "TR 03:30 PM - 04:45 PM", "BSN 1100", 0, "Spring 2025"],
    [15754, "Programming Concepts", "COP", "2510", "UG", "Schinnel Small", "TR 03:30 PM - 04:45 PM", "BSN 1100", 0, "Spring 2025"],
    [16690, "Robotics Process Automation", "CIS", "6930", "GR", "Marilynn Carol Hoeke", "TBA", "OFFT OFF", 0, "Spring 2025"],
    [16689, "Robotics Process Automation", "COP", "4931", "UG", "Marilynn Carol Hoeke", "TBA", "OFFT OFF", 0, "Spring 2025"],
    [15528, "Secure Coding", "CNT", "4419", "UG", "Jarred Ligatti", "MW 03:30 PM - 04:45 PM", "CHE 103", 0, "Spring 2025"],
    [15529, "Secure Coding", "CNT", "4419", "UG", "Xinming Ou", "MW 12:30 PM - 01:45 PM", "SOC 150", 0, "Spring 2025"],
    [15923, "Smart & Connected Health", "CIS", "4930", "UG", "John Templeton", "TR 11:00 AM - 12:15 PM", "SOC 152", 0, "Spring 2025"],
    [15924, "Smart & Connected Health", "CIS", "6930", "GR", "John Templeton", "TR 11:00 AM - 12:15 PM", "SOC 152", 0, "Spring 2025"],
    [19014, "Social Media Mining", "CAI", "5133", "GR", "Raiyan Abdul Baten", "TR 03:30 PM - 04:45 PM", "CIS 1048", 0, "Spring 2025"],
    [15930, "Social Media Mining", "CAP", "4773", "UG", "Raiyan Abdul Baten", "TR 03:30 PM - 04:45 PM", "CIS 1048", 0, "Spring 2025"],
    [15531, "Software Engineering", "CEN", "4020", "UG", "Zhao Han", "MW 08:00 AM - 09:15 AM", "CHE 100", 0, "Spring 2025"],
    [15532, "Software Engineering", "CEN", "4020", "UG", "Suey-Chyun Fang", "MW 03:30 PM - 04:45 PM", "BEH 104", 0, "Spring 2025"],
    [15543, "Software System Development", "COP", "4365", "UG", "Henry Jeanty", "TR 09:30 AM - 10:45 AM", "ENB 118", 0, "Spring 2025"],
    [15544, "Software System Development", "COP", "4365", "UG", "Henry Jeanty", "MW 09:30 AM - 10:45 AM", "ENB 118", 0, "Spring 2025"],
    [16692, "Sys Integration & Architecture", "CIS", "3433", "UG", "Marilynn Carol Hoeke", "MW 05:00 PM - 06:15 PM", "CHE 102", 0, "Spring 2025"],
    [20336, "Trustworthy AI Systems", "CIS", "6930", "GR", "Guangjing Wang", "MW 06:30 PM - 07:45 PM", "CHE 103", 0, "Spring 2025"],
    [15707, "Trustworthy Infrastructures", "CIS", "4212", "UG", "Attila Yavuz", "MW 02:00 PM - 03:15 PM", "CWY 109", 0, "Spring 2025"],
    [15706, "Trustworthy Infrastructures", "CIS", "6214", "GR", "Attila Yavuz", "MW 02:00 PM - 03:15 PM", "CWY 109", 0, "Spring 2025"],
    [20022, "User Experience Design", "CIS", "4930", "UG", "Holloman, Amanda K", "TR 03:30 PM - 04:45 PM", "CHE 103", 0, "Spring 2025"],
    [15936, "Wireless and Mobile Computing", "CIS", "4930", "UG", "Ren, Yili ", "MW 12:30 PM - 01:45 PM", "CHE 217", 0, "Spring 2025"],
    [15935, "Wireless and Mobile Computing", "CIS", "6930", "GR", "Ren, Yili ", "MW 12:30 PM - 01:45 PM", "CHE 217", 0, "Spring 2025"],
    [15933, "Sys Integration & Architecture", "CIS", "3433", "UG", "Hoeke, Marilynn Carol", "TBA", "TBAT TBA", 85, "Spring 2026"],
    [15930, "Robotics Process Automation", "COP", "4931", "UG", "Hoeke, Marilynn Carol", "TBA", "OFFT OFF", 66, "Spring 2026"],
    [15931, "Robotics Process Automation", "CIS", "6930", "GR", "Hoeke, Marilynn Carol", "TBA", "OFFT OFF", 0, "Spring 2026"],
    [20444, "Computing Circuits", "CDA", "4021", "UG", "Rahman, Taseef", "TR 12:30 PM - 01:45 PM", "EDU 347", 20, "Spring 2026"],
    [14881, "Computer Architecture", "CDA", "4205", "UG", "Rahman, Taseef", "MW 05:00 PM - 06:15 PM", "CHE 100", 70, "Spring 2026"],
    [16011, "Computer Architecture Lab", "CDA", "4205L", "UG", "Rahman, Taseef", "R 09:00 AM - 10:45 AM", "ENB 214", 34, "Spring 2026"],
    [16012, "Computer Architecture Lab", "CDA", "4205L", "UG", "Rahman, Taseef", "R 11:30 AM - 01:15 PM", "ENB 214", 35, "Spring 2026"],
    [14894, "Computer Logic & Design", "CDA", "3201", "UG", "Rahman, Taseef", "TR 02:00 PM - 03:15 PM", "ULH 101", 79, "Spring 2026"],
    [16013, "Computer Logic Design Lab", "CDA", "3201L", "UG", "Rahman, Taseef", "W 05:00 PM - 06:45 PM", "ENB 214", 34, "Spring 2026"],
    [16018, "Computer Logic Design Lab", "CDA", "3201L", "UG", "Rahman, Taseef", "F 08:00 AM - 09:45 AM", "ENB 214", 24, "Spring 2026"],
    [16019, "Computer Logic Design Lab", "CDA", "3201L", "UG", "Rahman, Taseef", "F 02:00 PM - 03:45 PM", "ENB 214", 35, "Spring 2026"],
    [16020, "Computer Logic Design Lab", "CDA", "3201L", "UG", "Rahman, Taseef", "T 11:30 AM - 01:15 PM", "ENB 214", 35, "Spring 2026"],
    [14923, "Advanced OOP in C++ for IT", "COP", "4931", "UG", "Jeanty, Henry H", "MW 12:30 PM - 01:45 PM", "ENB 118", 33, "Spring 2026"],
    [16023, "Computer Info Networks Lab", "CNT", "4104L", "UG", "Rauscher, Richard L", "F 01:00 PM - 02:50 PM", "NEC 300", 28, "Spring 2026"],
    [16024, "Computer Info Networks Lab", "CNT", "4104L", "UG", "Rauscher, Richard L", "R 01:00 PM - 02:50 PM", "NEC 300", 32, "Spring 2026"],
    [16025, "Computer Info Networks Lab", "CNT", "4104L", "UG", "Rauscher, Richard L", "F 09:00 AM - 10:50 AM", "NEC 300", 31, "Spring 2026"],
    [16026, "Computer Info Networks Lab", "CNT", "4104L", "UG", "Rauscher, Richard L", "R 11:00 AM - 12:50 PM", "NEC 300", 32, "Spring 2026"],
    [16027, "Computer Info Networks Lab", "CNT", "4104L", "UG", "Rauscher, Richard L", "R 09:00 AM - 10:50 AM", "NEC 300", 32, "Spring 2026"],
]


# --- private helpers ---------------------------------------------------------

def _parse_start_minutes(time_str):
    '''Return start time in minutes since midnight from a string like "MWF 09:00 AM - 09:50 AM", or None.'''
    m = re.search(r'(\d{1,2}):(\d{2})\s*(AM|PM)', time_str, re.IGNORECASE)
    if not m:
        return None
    h, mins, meridiem = int(m.group(1)), int(m.group(2)), m.group(3).upper()
    if meridiem == 'PM' and h != 12:
        h += 12
    elif meridiem == 'AM' and h == 12:
        h = 0
    return h * 60 + mins


def _parse_days(time_str):
    '''Return the day-letter string (e.g. "MWF") from a time string, or empty string.'''
    tokens = time_str.strip().split()
    if tokens and all(c in 'MTWRF' for c in tokens[0].upper()):
        return tokens[0].upper()
    return ''


def _match_course_code(code_query, subj, course_num):
    '''Match a query like "CDA" or "CDA 4205" against a course's subject code and number.
    The subject part uses substring matching; the number part uses substring matching too.'''
    if not code_query:
        return True
    parts = code_query.strip().upper().split(None, 1)
    subj_q = parts[0]
    num_q  = parts[1].strip() if len(parts) > 1 else ''
    if subj_q not in subj.upper():
        return False
    if num_q and num_q not in str(course_num):
        return False
    return True


#################################
### Functions used by the API ###
#################################

def search_courses(professor='', subject='', level='', course_code='',
                   semester='', days='', crn='') -> list:
    '''Search the course catalog. All parameters are optional; empty string means match all.

    days -- string of day letters (e.g. "MWF"); a course matches if it meets on ANY listed day.
    crn  -- exact CRN number to look up.
    Returns a list of dicts.'''
    results = []
    for row in COURSES:
        c_crn, name, subj, course_num, lvl, instr, time, room, enrolled, sem = row

        # exact CRN lookup
        if crn and str(c_crn) != crn.strip():
            continue

        # partial instructor match
        if professor and professor.lower() not in instr.lower():
            continue

        # subject filter — supports plain code ("CIS") or code+number ("CIS 4930")
        if subject and not _match_course_code(subject, subj, course_num):
            continue

        # exact level match
        if level and level.upper() != lvl:
            continue

        # course code match (e.g. "CDA" or "CDA 4205")
        if not _match_course_code(course_code, subj, course_num):
            continue

        # exact semester match
        if semester and semester != sem:
            continue

        # days filter — course must meet on at least one of the requested days
        if days:
            course_days = _parse_days(time)
            if not any(d in course_days for d in days.upper()):
                continue

        results.append({
            'crn':        c_crn,
            'name':       name,
            'subject':    subj,
            'course_num': course_num,
            'level':      lvl,
            'instructor': instr,
            'time':       time,
            'room':       room,
            'enrolled':   enrolled,
            'semester':   sem,
        })
    return results


def import_excel(filepath: str, semester: str) -> int:
    '''Import courses from a Bellini Classes Excel file into COURSES.
    Detects column layout automatically (F25 vs S25 vs S26 format).
    Returns the number of rows imported.'''
    from openpyxl import load_workbook
    wb = load_workbook(filepath, read_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0

    headers = [str(h).strip() if h else '' for h in rows[0]]

    def col(name):
        for h in headers:
            if name.lower() in h.lower().replace(' ', '_').replace(' ', ''):
                return headers.index(h)
        return None

    # find key columns by partial name match
    def find(keywords):
        for kw in keywords:
            for i, h in enumerate(headers):
                if kw.lower() in h.lower().replace(' ', ''):
                    return i
        return None

    idx_crn   = find(['crn'])
    idx_subj  = find(['subj'])
    idx_num   = find(['crse_numb', 'crsenumb', 'crse numb'])
    idx_title = find(['crse_title', 'crsetitle', 'crse title'])
    idx_levl  = find(['crse_levl', 'crselevl', 'crse levl'])
    idx_days  = find(['meeting_days', 'meetingdays', 'meeting days'])
    idx_times = find(['meeting_times1', 'meeting_times', 'meetingtimes'])
    idx_room  = find(['meeting_room', 'meetingroom'])
    idx_instr = find(['instructor'])
    idx_enrl  = find(['enrollment'])

    imported = 0
    for row in rows[1:]:
        try:
            crn   = int(row[idx_crn]) if idx_crn is not None and row[idx_crn] else None
            if not crn:
                continue
            subj  = str(row[idx_subj] or '').strip()
            num   = str(row[idx_num]  or '').strip()
            title = str(row[idx_title] or '').strip()
            levl  = str(row[idx_levl]  or '').strip()
            days  = str(row[idx_days]  or '').strip()
            times = str(row[idx_times] or '').strip()
            time  = f'{days} {times}'.strip() if days and days != 'None' else times
            room  = str(row[idx_room]  or '').strip()
            instr = str(row[idx_instr] or '').strip()
            enrl  = int(row[idx_enrl]) if idx_enrl is not None and row[idx_enrl] and str(row[idx_enrl]).isdigit() else 0

            # skip if CRN already loaded
            if any(c[0] == crn for c in COURSES):
                continue

            COURSES.append([crn, title, subj, num, levl, instr, time, room, enrl, semester])
            imported += 1
        except Exception:
            continue
    return imported


def add_class(class_data: dict) -> tuple[str, int]:
    '''Add a new class to the in-memory catalog.
    Expects a dict with keys: crn, name, subject, course_num, level, instructor, time, room, enrolled, semester.
    Returns a (message, status_code) tuple.'''

    try:
        crn = int(class_data.get('crn', 0))
    except (ValueError, TypeError):
        return ("Invalid CRN — must be a number.", 400)

    if not crn:
        return ("CRN is required.", 400)

    if any(c[0] == crn for c in COURSES):
        return (f"A course with CRN {crn} already exists.", 409)

    COURSES.append([
        crn,
        str(class_data.get('name',        '')).strip(),
        str(class_data.get('subject',     '')).strip(),
        str(class_data.get('course_num',  '')).strip(),
        str(class_data.get('level',       'UG')).strip(),
        str(class_data.get('instructor',  '')).strip(),
        str(class_data.get('time',        '')).strip(),
        str(class_data.get('room',        '')).strip(),
        int(class_data.get('enrolled',    0)),
        str(class_data.get('semester',    '')).strip(),
    ])
    return (f"Course {crn} added successfully.", 201)


def update_class(class_id: int, class_data: dict) -> tuple[str, int]:
    '''Update the course whose CRN matches class_id.
    Accepts a dict with any subset of: name, subject, course_num, level, instructor, time, room, enrolled, semester.
    Returns a (message, status_code) tuple.'''

    for i, row in enumerate(COURSES):
        if row[0] == class_id:
            crn, name, subj, num, lvl, instr, time, room, enrolled, sem = row
            COURSES[i] = [
                crn,
                str(class_data.get('name',        name)).strip(),
                str(class_data.get('subject',     subj)).strip(),
                str(class_data.get('course_num',  num)).strip(),
                str(class_data.get('level',       lvl)).strip(),
                str(class_data.get('instructor',  instr)).strip(),
                str(class_data.get('time',        time)).strip(),
                str(class_data.get('room',        room)).strip(),
                int(class_data.get('enrolled',    enrolled)),
                str(class_data.get('semester',    sem)).strip(),
            ]
            return (f"Course {class_id} updated successfully.", 200)

    return (f"No course found with CRN {class_id}.", 404)
